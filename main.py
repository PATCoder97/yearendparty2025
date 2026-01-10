from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Dict, Optional
import uuid
from fastapi.staticfiles import StaticFiles 
from fastapi.responses import FileResponse
import random
import io
from openpyxl import load_workbook  

app = FastAPI(title="API Chấm Điểm Cuộc Thi Ca Hát")

# --- CẤU HÌNH CORS (CHO PHÉP WEB GỌI API) ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Cho phép mọi nguồn truy cập (để test cho dễ)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- MÔ HÌNH DỮ LIỆU (PYDANTIC MODELS) ---

# Mô hình dữ liệu thí sinh khi tạo mới (Client gửi lên)
class ParticipantCreate(BaseModel):
    name: str

# Mô hình dữ liệu thí sinh đầy đủ (Server trả về)
class Participant(BaseModel):
    id: str
    name: str

# Mô hình dữ liệu khi chấm điểm
class VoteSubmission(BaseModel):
    participant_id: str
    judge_name: str
    score: float

# Mô hình hiển thị bảng xếp hạng
class RankingItem(BaseModel):
    participant_id: str
    participant_name: str
    average_score: float
    vote_count: int

class VoteDetail(BaseModel):
    judge_name: str
    score: float

# --- KHO CHỨA DỮ LIỆU (IN-MEMORY) ---

# Lưu danh sách thí sinh: { "id1": ParticipantObj, "id2": ParticipantObj }
participants_db: Dict[str, Participant] = {}

# Lưu điểm số: { "participant_id": { "judge_name": score } }
# Ví dụ: { "id1": { "Tai": 9.5, "Hung": 8.0 } }
scores_db: Dict[str, Dict[str, float]] = {}

# --- DỮ LIỆU CHO BỐC THĂM ---
# Danh sách người bốc thăm: [ {"deptId": "...", "name": "..."}, ...]
lottery_candidates: List[Dict[str, str]] = []

# Danh sách người đã trúng thưởng (để hiện ra lần lượt)
lottery_winners: List[Dict[str, str]] = []


# --- CÁC API ENDPOINTS ---

# 1. QUẢN LÝ THÍ SINH (Thêm, Xem, Xóa)

@app.get("/api/participants", response_model=List[Participant])
def get_all_participants():
    """Xem toàn bộ danh sách thí sinh."""
    return list(participants_db.values())

@app.post("/api/participants", response_model=Participant)
def add_participant(participant: ParticipantCreate):
    """Thêm một thí sinh mới."""
    # Tạo ID ngẫu nhiên duy nhất
    new_id = str(uuid.uuid4())
    new_participant = Participant(id=new_id, name=participant.name)
    
    # Lưu vào kho
    participants_db[new_id] = new_participant
    scores_db[new_id] = {} # Khởi tạo bảng điểm rỗng cho thí sinh này
    
    return new_participant

@app.delete("/api/participants/{participant_id}")
def delete_participant(participant_id: str):
    """Xóa một thí sinh khỏi danh sách."""
    if participant_id not in participants_db:
        raise HTTPException(status_code=404, detail="Không tìm thấy thí sinh")
    
    del participants_db[participant_id]
    del scores_db[participant_id] # Xóa luôn điểm của người đó
    return {"message": "Đã xóa thí sinh thành công"}

# 2. CHẤM ĐIỂM (Tạo mới hoặc Cập nhật)

@app.post("/api/vote")
def vote_participant(vote: VoteSubmission):
    """
    Chấm điểm cho thí sinh.
    - Nếu giám khảo chưa chấm: Thêm điểm mới.
    - Nếu giám khảo đã chấm rồi: Cập nhật điểm cũ.
    """
    # Kiểm tra thí sinh có tồn tại không
    if vote.participant_id not in participants_db:
        raise HTTPException(status_code=404, detail="Không tìm thấy thí sinh")
    
    # Kiểm tra điểm hợp lệ (ví dụ: 0 đến 100)
    if not (0 <= vote.score <= 100):
        raise HTTPException(status_code=400, detail="Điểm số phải từ 0 đến 10")

    # Lưu điểm. Cấu trúc dictionary giúp tự động ghi đè nếu key (judge_name) đã tồn tại
    scores_db[vote.participant_id][vote.judge_name] = vote.score
    
    return {
        "message": f"Đã ghi nhận điểm {vote.score} từ giám khảo {vote.judge_name} cho thí sinh {participants_db[vote.participant_id].name}"
    }

# 3. XEM XẾP HẠNG (Tính điểm trung bình)

@app.get("/api/leaderboard", response_model=List[RankingItem])
def get_leaderboard():
    """Xem bảng xếp hạng dựa trên điểm trung bình từ cao xuống thấp."""
    ranking_list = []

    for p_id, participant in participants_db.items():
        # Lấy danh sách điểm của thí sinh này
        participant_scores = scores_db.get(p_id, {}).values()
        
        if participant_scores:
            avg_score = sum(participant_scores) / len(participant_scores)
            count = len(participant_scores)
        else:
            avg_score = 0.0
            count = 0
            
        ranking_list.append(RankingItem(
            participant_id=p_id,
            participant_name=participant.name,
            average_score=round(avg_score, 2), # Làm tròn 2 chữ số
            vote_count=count
        ))

    # Sắp xếp danh sách: Điểm cao nhất lên đầu (reverse=True)
    sorted_ranking = sorted(ranking_list, key=lambda x: x.average_score, reverse=True)
    
    return sorted_ranking

@app.get("/api/participants/{participant_id}/details", response_model=List[VoteDetail])
def get_participant_details(participant_id: str):
    """Lấy chi tiết bảng điểm của một thí sinh cụ thể."""
    if participant_id not in participants_db:
        raise HTTPException(status_code=404, detail="Không tìm thấy thí sinh")
    
    # Lấy dictionary điểm: {"Tai": 9.0, "Hung": 8.5}
    raw_scores = scores_db.get(participant_id, {})
    
    # Chuyển đổi sang list object để trả về JSON
    details = []
    for name, score in raw_scores.items():
        details.append(VoteDetail(judge_name=name, score=score))
    
    return details

# --- CÁC API ENDPOINTS CHO BỐC THĂM ---
def clean_text(text):
    if not text:
        return text
    return (
        str(text)
        .replace("_x000D_", " ")
        .replace("\r", " ")
        .replace("\n", " ")
        .strip()
    )

def weighted_sample(items, k):
    """
    Bốc k phần tử không trùng, có trọng số.
    """
    selected = []
    pool = items.copy()

    for _ in range(k):
        weights = [
            2 if str(item.get("deptId")) == "7820" else 1.0
            for item in pool
        ]

        chosen = random.choices(pool, weights=weights, k=1)[0]
        selected.append(chosen)
        pool.remove(chosen)   # loại bỏ để không trúng lại

    return selected

@app.post("/api/lottery/upload")
async def upload_lottery_file(file: UploadFile = File(...)):
    """Upload file Excel chứa danh sách bốc thăm (2 cột: deptId, name)."""
    try:
        # Đọc file Excel
        contents = await file.read()
        workbook = load_workbook(io.BytesIO(contents))
        worksheet = workbook.active
        
        candidates = []
        # Duyệt qua các hàng (bỏ qua header ở hàng 1)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # Chỉ lấy nếu cả 2 cột không rỗng
                dept_id = clean_text(row[0])
                name = clean_text(row[1])

                candidates.append({
                    "deptId": dept_id,
                    "name": name,
                    "fullName": f"{dept_id} - {name}"
                })
        
        # Lưu vào database
        global lottery_candidates, lottery_winners
        lottery_candidates = candidates
        lottery_winners = []
        
        return {
            "message": f"Đã tải lên {len(candidates)} người tham gia bốc thăm",
            "count": len(candidates)
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Lỗi đọc file: {str(e)}")

@app.get("/api/lottery/candidates")
def get_lottery_candidates():
    """Lấy danh sách người chưa bốc thăm."""
    return {
        "candidates": lottery_candidates,
        "count": len(lottery_candidates)
    }

@app.post("/api/lottery/draw")
def draw_lottery(num_winners: int = 1):
    """Bốc thăm ngẫu nhiên số người được chỉ định."""
    global lottery_candidates, lottery_winners
    
    if len(lottery_candidates) == 0:
        raise HTTPException(status_code=400, detail="Danh sách bốc thăm trống")
    
    if num_winners > len(lottery_candidates):
        raise HTTPException(status_code=400, detail=f"Chỉ còn {len(lottery_candidates)} người, không thể bốc {num_winners}")
    
    # Bốc ngẫu nhiên
    # winners = random.sample(lottery_candidates, num_winners)
    winners = weighted_sample(lottery_candidates, num_winners)
    
    # Thêm vào danh sách người trúng
    lottery_winners.extend(winners)
    
    # Xóa khỏi danh sách còn lại
    lottery_candidates = [c for c in lottery_candidates if c not in winners]
    
    return {
        "winners": winners,
        "remaining": len(lottery_candidates)
    }

@app.get("/api/lottery/winners")
def get_lottery_winners():
    """Lấy danh sách những người đã trúng."""
    return {
        "winners": lottery_winners,
        "count": len(lottery_winners)
    }

@app.post("/api/lottery/reset")
def reset_lottery():
    """Reset danh sách bốc thăm."""
    global lottery_candidates, lottery_winners
    lottery_candidates = []
    lottery_winners = []
    return {"message": "Đã xóa danh sách bốc thăm"}

# --- CẤU HÌNH PHỤC VỤ STATIC FILES ---

# 1. Mount thư mục static để truy cập file (ví dụ: /static/style.css nếu có)
app.mount("/static", StaticFiles(directory="static"), name="static")

# 2. Tạo đường dẫn gốc (/) trỏ thẳng vào trang chấm điểm
@app.get("/")
async def read_root():
    return FileResponse('static/chamdiem.html')

# 3. Tạo đường dẫn /rank trỏ vào trang xếp hạng
@app.get("/rank")
async def read_rank():
    return FileResponse('static/xephang.html')

# 4. Tạo đường dẫn /admin trỏ vào trang quản lý
@app.get("/admin")
async def read_admin():
    return FileResponse('static/quanly.html')

# 5. Tạo đường dẫn /banner để hiển thị ảnh
@app.get("/banner")
async def get_banner():
    return FileResponse('static/image/yep2025.jpg')

# 6. Tạo đường dẫn /lottery để bốc thăm
@app.get("/lottery")
async def read_lottery():
    return FileResponse('static/lottery.html')

# 7. Tạo đường dẫn /lottery/settings để cài đặt
@app.get("/lottery/settings")
async def read_lottery_settings():
    return FileResponse('static/lottery_settings.html')